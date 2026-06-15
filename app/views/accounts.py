"""账号管理 API"""

import os
import csv
import io
import logging

from flask import Blueprint, request, jsonify, make_response, session

from .. import requires_auth, require_role
from ..models.db import get_db, get_db_dict
from datetime import datetime

log = logging.getLogger(__name__)
accounts_bp = Blueprint("accounts", __name__)


@accounts_bp.route("/api/accounts", methods=["GET"])
@requires_auth
def api_accounts():
    try:
        page = request.args.get("page", 1, type=int)
        limit = request.args.get("limit", 20, type=int)
        search = request.args.get("search", "")

        conn = get_db_dict()
        cur = conn.cursor()
        if search:
            like = f"%{search}%"
            cur.execute(
                "SELECT COUNT(*) as c FROM tn_accounts WHERE username LIKE ? OR phone_number LIKE ?",
                (like, like),
            )
        else:
            cur.execute("SELECT COUNT(*) as c FROM tn_accounts")
        total = cur.fetchone()["c"]

        offset = (page - 1) * limit
        if search:
            like = f"%{search}%"
            cur.execute(
                """SELECT a.*, p.ip as bind_ip, s.name as salesman_name
                FROM tn_accounts a
                LEFT JOIN tn_account_ip ai ON a.id = ai.account_id
                LEFT JOIN tn_ip_pool p ON ai.ip_id = p.id
                LEFT JOIN tn_salesman s ON a.salesman_id = s.id
                WHERE a.username LIKE ? OR a.phone_number LIKE ?
                ORDER BY a.id DESC LIMIT ? OFFSET ?""",
                (like, like, limit, offset),
            )
        else:
            cur.execute(
                """SELECT a.*, p.ip as bind_ip, s.name as salesman_name
                FROM tn_accounts a
                LEFT JOIN tn_account_ip ai ON a.id = ai.account_id
                LEFT JOIN tn_ip_pool p ON ai.ip_id = p.id
                LEFT JOIN tn_salesman s ON a.salesman_id = s.id
                ORDER BY a.id DESC LIMIT ? OFFSET ?""",
                (limit, offset),
            )
        rows = cur.fetchall()
        conn.close()
        return jsonify({"code": 0, "msg": "", "count": total, "data": [dict(r) for r in rows]})
    except Exception as e:
        log.error("api_accounts error: %s", e)
        return jsonify({"code": 1, "msg": str(e), "count": 0, "data": []})


@accounts_bp.route("/api/accounts", methods=["POST"])
@requires_auth
def api_add_account():
    try:
        data = request.json
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO tn_accounts (username, password, sid, token, status)
               VALUES (?, ?, ?, ?, 1)""",
            (data.get("username"), data.get("password"), data.get("sid"), data.get("token")),
        )
        conn.commit()
        conn.close()
        return jsonify({"code": 0, "msg": "添加成功"})
    except Exception as e:
        log.error("api_add_account error: %s", e)
        return jsonify({"code": 1, "msg": str(e)})


@accounts_bp.route("/api/accounts/<int:aid>", methods=["GET"])
@requires_auth
def api_account_detail(aid):
    try:
        conn = get_db_dict()
        cur = conn.cursor()
        cur.execute(
            """SELECT a.*, p.ip as bind_ip
            FROM tn_accounts a
            LEFT JOIN tn_account_ip ai ON a.id = ai.account_id
            LEFT JOIN tn_ip_pool p ON ai.ip_id = p.id
            WHERE a.id=?""",
            (aid,),
        )
        row = cur.fetchone()
        conn.close()
        if row:
            return jsonify({"code": 0, "msg": "", "data": dict(row)})
        return jsonify({"code": 1, "msg": "账号不存在"})
    except Exception as e:
        log.error("api_account_detail error: %s", e)
        return jsonify({"code": 1, "msg": str(e)})


@accounts_bp.route("/api/accounts/update", methods=["POST"])
@requires_auth
def api_update_account():
    try:
        data = dict(request.form) if request.form else (request.json or {})
        if not data.get("id"):
            return jsonify({"code": 1, "msg": "缺少ID"})

        conn = get_db()
        cur = conn.cursor()
        fields = []
        values = []
        skip_pwd = not data.get("password")
        all_fields = [
            "username",
            "phone_number",
            "email",
            "sid",
            "token",
            "user_agent",
            "px_auth",
            "device_fp",
            "idfa",
            "client_id",
            "proxy",
            "status",
            "health_score",
        ]
        if not skip_pwd:
            all_fields.insert(1, "password")

        for k in all_fields:
            v = data.get(k)
            if v is not None:
                fields.append(f"{k}=?")
                values.append(v)

        if fields:
            values.append(data["id"])
            cur.execute(f"UPDATE tn_accounts SET {','.join(fields)} WHERE id=?", values)
            conn.commit()
        conn.close()
        return jsonify({"code": 0, "msg": "保存成功"})
    except Exception as e:
        log.error("api_update_account error: %s", e)
        return jsonify({"code": 1, "msg": str(e)})


@accounts_bp.route("/api/accounts/delete", methods=["POST"])
@requires_auth
def api_delete_account():
    try:
        aid = request.form.get("id") or request.json.get("id")
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT ip_id FROM tn_account_ip WHERE account_id=?", (aid,))
        ip_row = cur.fetchone()
        if ip_row:
            cur.execute("DELETE FROM tn_account_ip WHERE account_id=?", (aid,))
            cur.execute("UPDATE tn_ip_pool SET status=1 WHERE id=?", (ip_row[0],))
        cur.execute("DELETE FROM tn_accounts WHERE id=?", (aid,))
        conn.commit()
        conn.close()
        return jsonify({"code": 0, "msg": "删除成功，IP已回收"})
    except Exception as e:
        log.error("api_delete_account error: %s", e)
        return jsonify({"code": 1, "msg": str(e)})


@accounts_bp.route("/api/accounts/export", methods=["GET"])
@requires_auth
def api_export_accounts():
    try:
        fmt = request.args.get("format", "csv")
        scope = request.args.get("scope", "all")
        mode = request.args.get("mode", "internal")
        ids_str = request.args.get("ids", "")

        conn = get_db_dict()
        cur = conn.cursor()

        query = """SELECT a.id, a.username, a.password, a.phone_number, a.email, a.status,
                          a.health_score, a.sid, a.token, a.proxy, a.user_agent, a.device_fp,
                          a.idfa, a.client_id, a.px_auth, a.last_used_at,
                          p.ip as bind_ip, p.port as bind_port, p.proxy_user, p.proxy_pwd,
                          p.area as ip_area, p.ip_type, p.id as ip_row_id
                   FROM tn_accounts a
                   LEFT JOIN tn_account_ip ai ON a.id = ai.account_id
                   LEFT JOIN tn_ip_pool p ON ai.ip_id = p.id"""
        conditions = []
        params = []

        if scope == "active":
            conditions.append("a.status = 1")
        elif scope == "selected" and ids_str:
            id_list = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]
            placeholders = ",".join(["?"] * len(id_list))
            conditions.append(f"a.id IN ({placeholders})")
            params.extend(id_list)

        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY a.id"

        cur.execute(query, params)
        rows = [dict(r) for r in cur.fetchall()]

        if mode == "external":
            conn2 = get_db()
            cur2 = conn2.cursor()
            unbound_ips = [r["ip_row_id"] for r in rows if r.get("ip_row_id")]
            for row in rows:
                if row.get("ip_row_id"):
                    cur2.execute("DELETE FROM tn_account_ip WHERE account_id=?", (row["id"],))
            if unbound_ips:
                ph = ",".join(["?"] * len(unbound_ips))
                cur2.execute(f"UPDATE tn_ip_pool SET status=1 WHERE id IN ({ph})", unbound_ips)
            conn2.commit()
            conn2.close()

        conn.close()

        def proxy_str(row):
            if not row.get("bind_ip"):
                return ""
            auth = ""
            if row.get("proxy_user") and row.get("proxy_pwd"):
                auth = f"{row['proxy_user']}:{row['proxy_pwd']}@"
            return f"http://{auth}{row['bind_ip']}:{row.get('bind_port', '')}"

        if fmt == "csv":
            si = io.StringIO()
            writer = csv.writer(si)
            writer.writerow(
                [
                    "ID", "用户名", "手机号", "邮箱", "状态", "健康度",
                    "User-Agent", "设备指纹",
                    "IDFA", "ClientID", "绑定IP", "IP地区", "IP类型", "代理连接串",
                ]
            )
            status_map = {1: "正常", 0: "禁用", 2: "异常"}
            for r in rows:
                writer.writerow(
                    [
                        r["id"], r["username"], r.get("phone_number", ""),
                        r.get("email", ""), status_map.get(r.get("status"), ""), r.get("health_score", ""),
                        r.get("user_agent", ""), r.get("device_fp", ""), r.get("idfa", ""),
                        r.get("client_id", ""),
                        r.get("bind_ip", ""), r.get("ip_area", ""), r.get("ip_type", ""),
                        proxy_str(r),
                    ]
                )
            output = si.getvalue()
            mimetype, ext = "text/csv", "csv"
        else:
            lines = [f"{r['username']}----{r.get('phone_number', '')}----{proxy_str(r)}" for r in rows]
            output = "\n".join(lines)
            mimetype, ext = "text/plain", "txt"

        resp = make_response(output)
        resp.headers["Content-Type"] = f"{mimetype}; charset=utf-8"
        resp.headers["Content-Disposition"] = f"attachment; filename=accounts_export.{ext}"
        return resp
    except Exception as e:
        log.error("api_export_accounts error: %s", e)
        return jsonify({"code": 1, "msg": str(e)})


@accounts_bp.route("/api/accounts/batch_import", methods=["POST"])
@requires_auth
def api_batch_import_accounts():
    import json as json_mod

    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"code": 1, "msg": "未选择上传文件"})

        filename = file.filename or ""
        ext = os.path.splitext(filename)[1].lower()
        rows_data = []
        try:
            if ext == ".csv":
                text = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    rows_data.append(row)
            elif ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                except ImportError:
                    return jsonify({"code": 1, "msg": "需要 openpyxl 库，请执行: pip install openpyxl"})
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows_data.append(dict(zip(headers, row)))
                wb.close()
            elif ext == ".jsonl":
                text = file.read().decode("utf-8")
                for line in text.strip().splitlines():
                    line = line.strip()
                    if line:
                        rows_data.append(json_mod.loads(line))
            else:
                return jsonify({"code": 1, "msg": f"不支持的格式: {ext}，仅支持 csv/xlsx/jsonl"})
        except Exception as e:
            return jsonify({"code": 1, "msg": f"文件解析失败：{str(e)}"})

        if not rows_data:
            return jsonify({"code": 1, "msg": "文件内容为空"})

        conn = get_db()
        cur = conn.cursor()
        succ = 0
        fail = 0
        errors = []

        for i, row in enumerate(rows_data):
            try:
                username = str(row.get("username") or row.get("Username") or "").strip()
                if not username:
                    continue
                phone = str(row.get("phone_number") or row.get("phone") or row.get("Phone") or "").strip()
                cur.execute("SELECT id FROM tn_accounts WHERE username=? OR phone_number=?", (username, phone))
                if cur.fetchone():
                    fail += 1
                    errors.append(f"第{i+2}行: {username} 已存在")
                    continue

                cur.execute(
                    """INSERT INTO tn_accounts (
                        username, password, sid, token, phone_number, email,
                        idfa, user_agent, px_auth, device_fp,
                        os_version, client_id, proxy, status, health_score
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        username,
                        str(row.get("password") or row.get("Password") or ""),
                        str(row.get("sid") or row.get("SID") or ""),
                        str(row.get("token") or row.get("Token") or ""),
                        phone,
                        str(row.get("email") or row.get("Email") or ""),
                        str(row.get("idfa") or row.get("IDFA") or ""),
                        str(row.get("user_agent") or row.get("User-Agent") or ""),
                        str(row.get("px_auth") or row.get("X-PX-AUTHORIZATION") or row.get("token") or ""),
                        str(row.get("device_fp") or row.get("X-PX-DEVICE-FP") or ""),
                        str(row.get("os_version") or row.get("X-PX-OS-VERSION") or ""),
                        str(row.get("client_id") or row.get("clientId") or ""),
                        str(row.get("proxy") or ""),
                        int(row.get("status", 2)) if row.get("status") is not None else 2,
                        int(row.get("health_score", 0)) if row.get("health_score") is not None else 0,
                    ),
                )
                succ += 1
            except Exception as e:
                fail += 1
                errors.append(f"第{i+2}行: {e}")

        conn.commit()
        conn.close()

        msg = f"导入完成：成功 {succ} 条，跳过/失败 {fail} 条"
        if errors and len(errors) <= 5:
            msg += "\n" + "; ".join(errors)
        return jsonify({"code": 0, "msg": msg})
    except Exception as e:
        log.error("api_batch_import error: %s", e)
        return jsonify({"code": 1, "msg": str(e)})
